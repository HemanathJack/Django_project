#!/usr/bin/python3
"""
"""

from Application.py_files.common_lib.robot_test_case_op import robot_test_case_op
import os
import openpyxl

WORKSHEET_NAME = 'Test Cases'

PATH = "Application/static/resources/iot/rellab_iot/tests/DTU/"

class iot_testcase_op(robot_test_case_op):
	def __init__(self):
		super().__init__()
		self.common_robot = robot_test_case_op()
		self.iot_test_case = dict()
		self.testcase_owner = dict()
		self.testcase_documentation = dict()
		self.simulator_testcase = dict()

	def add_simulator_name(simu,content):
		"""Modify excel file in static location with respect to selected simulator.

		Args:
			simu (str): Simulator ID.
			content (dict): Catagory and test case title.

		Returns:
			xlsx file with selected simulator test case information. 
		"""
		iot_database = openpyxl.load_workbook('Application/static/iot.xlsx')
		iot = iot_database[WORKSHEET_NAME]
		iot.delete_rows(1, iot.max_row)

		row_start = 3

		iot.cell(row=row_start, column=2).value = str(simu)
		for job, cases in content.items():
			iot.cell(row=row_start, column=3).value = str(job)
			for case in cases:
				iot.cell(row=row_start, column=4).value = str(case)
				row_start += 1
			row_start += 1
		iot_database.save('Application/static/iot.xlsx')
		return iot_database

	def categorise_iot_testcase(self):
		"""Categorise IOT Repository test case by category.

		Returns:
			Catagorise with test case name (dict). | key:str, value:list
		"""
		self.create_defined_categories()
		_suites = self.common_robot.read_suits_and_return_testcase_with_tags(PATH)
		for suite, test_cases in _suites.items():
			for case, tags in test_cases.items():
				self.get_test_case_catagory(case, suite, tags)
		return self.iot_test_case

	def catagorise_iot_simulator_testcase(self):
		"""Catagorise IOT test cases with respect to simulators
		
		Returns:
			Catagorise test case with iot simulator (dict). | key:str, value:list
		"""
		self.create_iot_simualtor_catagories()
		_suites = self.common_robot.read_suits_and_return_testcase_with_tags(PATH)
		for suite, testcase_items in _suites.items():
			for testcase, tags in testcase_items.items():
				if 'esc' in suite or 'anyesc' in suite:
					self.simulator_testcase['SIMU232'].append(testcase)
				else:
					if all(['SIMU' not in _tag for _tag in tags]):
						for _simu in self.simulator_testcase:
							self.simulator_testcase[_simu].append(testcase)
					for _simu in self.simulator_testcase:
						if _simu in tags:
							self.simulator_testcase[_simu].append(testcase)
		return self.simulator_testcase

	def convert_string_to_suite_form(self, string):
		"""Convert string to iot test suite name format

		Args:
			string (str): String to convert into format.
		
		Example:
			Cloud View -> cloud_view
		
		Returns:
			formated string (str)
		"""
		return string.lower().replace(" ", '_')

	def create_defined_categories(self):
		"""Defined categories for IOT
		"""
		_catagories = ['Anyescalator', 'Cloud View', 'Device', 'Others', 'Software Update', 'Floor Statistics', 'Edge Computing',
			'Security Monitoring', 'Provision' , 'Remote Call' , 'Network' , 'Power' , 'Controller' , 'IBM Rollback', 'Escalator Edge Computing',
			'Escalator' , 'AWS Migration']
		for catagory in _catagories:
			self.iot_test_case[catagory] = list()
	
	def create_iot_simualtor_catagories(self):
		"""Create IOT Simualtor wise dict
		"""
		_simulatos = ['SIMU01', 'SIMU06', 'SIMU74', 'SIMU44', 'SIMU144', 'SIMU153', 'SIMU232', 'SIMU358', 'SIMU362', 'SIMU376']
		for simu in _simulatos:
			self.simulator_testcase[simu] = list()

	def fetch_documentation_of_test_cases(self):
		"""Fetch documentation of all test cases if available

		Returns:
			Dictionary of all test case from repository with documentation (dict). | key:str, value:str
		"""
		_suites = self.common_robot.read_suits_and_return_testcase_with_tags(PATH)
		for _file in os.listdir(PATH):
			if all(_file_name != _file for _file_name in self.ExcludedFilenames): 
				robot_file = open(PATH+_file, 'r+')
				for test_case in _suites[_file]:
					robot_file.seek(0)
					self.testcase_documentation[test_case] = self.get_documentation_of_testcase(robot_file, test_case)
		return self.testcase_documentation

	def fetch_gid_number_of_case(self):
		"""Fetch GID number of IOT test case.

		Returns:
			Dictionary of test case with gid number. (dict) | key:str, value:str
		"""
		test_case_id = dict()
		_suites = self.common_robot.read_suits_and_return_testcase_with_tags(PATH)
		for suite, test_cases in _suites.items():
			for case, tags in test_cases.items():
				for _tag in tags:
					if 'gid' in _tag:
						test_case_id[case] = _tag
						break
					test_case_id[case] = None
		return test_case_id

	def fetch_owners_of_test_cases(self):
		"""Fetch owner of test cases

		Returns:
			Dictionary of test case name with owner name. (dict) | key:str, value:str
		"""
		_suites = self.common_robot.read_suits_and_return_testcase_with_tags(PATH)
		for suite, test_cases in _suites.items():
			for case, tags in test_cases.items():
				self.testcase_owner[case] = self.get_owner_of_test_case(tags)
		return self.testcase_owner

	def get_counts_of_cases(self, case_collection):
		"""Get total, running, bug open and not ready counts of test cases

		Args:
			case_collection (dict): Dictionary of case collection.
		
		Returns:
			Counts of respective topics (dict).
		"""
		returnable = dict()
		_excludables = self.segregate_bug_and_not_ready_cases()
		returnable['total'] = 0
		returnable['bug'] = 0
		returnable['not_ready'] = 0
		for catagory, cases in case_collection.items():
			for case in cases:
				returnable['total'] += 1
				if case in _excludables['not_ready']:
					returnable['not_ready'] += 1
				elif case in _excludables['bug_open']:
					returnable['bug'] +=1
		returnable['running'] = returnable['total'] - (returnable['bug'] + returnable['not_ready'])
		return returnable

	def get_test_case_catagory(self, test_case, suite_name, tags):
		"""Get test case catagory and append test case name in the correct list.

		Args:
			test_case (str): Test case name.
			suite_name (str): Test suite name.
			tags (list): Tags for test case inc force tags.
		"""
		others = ['pmq', 'hax', 'iot_apis']
		if 'edge' in suite_name:
			if 'esc' in suite_name:
				self.iot_test_case['Escalator Edge Computing'].append(test_case)
				return
			self.iot_test_case['Edge Computing'].append(test_case)
			return
		if 'esc_move' in suite_name:
			self.iot_test_case['Escalator'].append(test_case)
			return
		if 'security_monitoring' in suite_name:
			self.iot_test_case['Security Monitoring'].append(test_case)
			return
		if 'provision' in suite_name:
			self.iot_test_case['Provision'].append(test_case)
			return
		if 'migration' in suite_name:
			self.iot_test_case['AWS Migration'].append(test_case)
			return
		if 'device_controller' in suite_name:
			self.iot_test_case['Controller'].append(test_case)
			return
		if 'cloud_view' in suite_name:
			self.iot_test_case['Cloud View'].append(test_case)
			return
		if 'device_power' in suite_name:
			self.iot_test_case['Power'].append(test_case)
			return
		if 'floor_statistics' in suite_name:
			self.iot_test_case['Floor Statistics'].append(test_case)
			return
		if 'software_update' in suite_name:
			self.iot_test_case['Software Update'].append(test_case)
			return
		if 'remote_call' in suite_name:
			self.iot_test_case['Remote Call'].append(test_case)
			return
		if 'rollback' in suite_name:
			self.iot_test_case['IBM Rollback'].append(test_case)
			return
		if 'anyescalator' in suite_name:
			self.iot_test_case['Anyescalator'].append(test_case)
			return
		if 'connectivity_monitoring' in suite_name or 'device_network' in suite_name:
			self.iot_test_case['Network'].append(test_case)
			return
		if any([_any in suite_name for _any in others]):
			self.iot_test_case['Others'].append(test_case)
			return
		self.iot_test_case['Device'].append(test_case)
	
	def get_simulator_type(self, simulator):
		"""Analyse simulator ID from defined dict and returns elevator or escalator 
		
		Args:
			simulator (str): Simulator ID to verify.
		"""
		_type = {
			'elevator' : ['SIMU01', 'SIMU06', 'SIMU74', 'SIMU144', 'SIMU153', 'SIMU358', 'SIMU376'],
			'escalator' : ['SIMU232', 'SIMU362']
		}
		for type, simus in _type.items():
			if simulator in simus:
				return type

	def organise_simuwise_with_catagories(self, simulator):
		"""Organise catagorise cases to simulator wise catagories
		
		Args:
			simulator (str): Simulator name.
		
		Returns:
			Dictionary of simulator wise cases in catagorised order (dict).
		"""
		_catagorised_cases = self.categorise_iot_testcase()
		_simulator_case = self.catagorise_iot_simulator_testcase()[simulator]
		returnable = dict()
		for catagory, cases in _catagorised_cases.items():
			returnable[catagory] = list()
			for _simucase in _simulator_case:
				if _simucase in cases:
					returnable[catagory].append(_simucase)
		none_keys = []
		for catagory in returnable:
			if not returnable[catagory]:
				none_keys.append(catagory)
		if none_keys:
			for _item in none_keys:
				del returnable[_item]
		if self.get_simulator_type(simulator) == 'elevator':
			_esc_catagories = ['Escalator', 'Anyescalator', 'Escalator Edge Computing']
			for _rem in _esc_catagories:
				try:
					del returnable[_rem]
				except:
					pass
		return returnable

	def return_available_simulators(self):
		"""Return available simulators in list form.

		Returns:
			list of available IOT simulators (list).
		"""
		self.create_iot_simualtor_catagories()
		return self.simulator_testcase.keys()

	def segregate_bug_and_not_ready_cases(self):
		"""Find bug open cases and not ready cases from the repository

		Returuns:
			Dictionary of not ready and bug openn cases with valus in list. (dict)
		"""
		returnable = {'not_ready': [], 'bug_open': []}
		_suites = self.common_robot.read_suits_and_return_testcase_with_tags(PATH)
		for suite, testcase_items in _suites.items():
			for testcase, tags in testcase_items.items():

				if 'not_ready' in tags:
					returnable['not_ready'].append(testcase)
				elif any([_tag.endswith('_open') for _tag in tags]):
					returnable['bug_open'].append(testcase)
		return returnable
	
	def simuwise_bug_and_not_ready_cases(self):
		"""Simulator wise bug and not ready cases from the repository
		
		Returns:
			Dictionary of simulatorwise not ready and bug open cases. (dict)
		"""
		returnable = dict()
		_overall_bugs_nt_rdy_cases = self.segregate_bug_and_not_ready_cases()
		
		for simu, cases in self.catagorise_iot_simulator_testcase().items():
			returnable[simu] = {'not_ready' : [], 'bug_open': []}
			for case in cases:
				if case in _overall_bugs_nt_rdy_cases['not_ready']:
					returnable[simu]['not_ready'].append(case)
				elif case in _overall_bugs_nt_rdy_cases['bug_open']:
					returnable[simu]['bug_open'].append(case)
		return returnable
