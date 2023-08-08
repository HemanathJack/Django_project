#!/usr/bin/python3
"""Operation functions for competency matrix.
"""

from Application.models import *
import openpyxl
from datetime import datetime, timedelta
import pytz
import tzlocal

WORKSHEET_NAME = 'Competency'

def add_update_competency(updates, user):
    """Documentation needs to be added

    Args:
        updates (type): needs to be updated.
        user (type): needs to be updated.
    """
    database = openpyxl.load_workbook('Application/static/db.xlsx')
    db = database[WORKSHEET_NAME]

    for i in range(4, db.max_column+1):
        if user == db.cell(row=16, column=i).value:
            user_column = i

    for i in range(18, db.max_row+1):
        for m, p in updates.items():
            if m == db.cell(row=i, column=2).value:
                for j in range(i, db.max_row+1):
                    for pr, val in p.items():
                        if pr == db.cell(row=j, column=3).value:
                            db.cell(row=j, column=user_column).value = val[0]
                            db.cell(row=j, column=user_column+1).value = val[1]

    database.save('Application/static/db.xlsx')

def extract_data_from_excel(individual):
    """Extract data of an individual user from excel db.

    Args:
        individual (str): Firstname + Lastname of user from database.

    Returns:
        compencty[user] (dict) : competency of specific user from excel.
    """
    database = openpyxl.load_workbook('Application/static/db.xlsx')
    db = database[WORKSHEET_NAME]

    competency = {}
    start = 18 #start row where the modules and products are defined.

    # creating nested dictionary with users as key
    for user in range(4, db.max_column+1):
        if db.cell(row=16, column=user).value == individual:
            competency[db.cell(row=16, column=user).value] = {}
            for mod in range(start, db.max_row+1):
                if db.cell(row=mod, column=2).value != None and db.cell(row=mod, column=2).value not in competency[db.cell(row=16, column=user).value]:
                    competency[db.cell(row=16, column=user).value][db.cell(row=mod, column=2).value] = {}
            for prod in range(start, db.max_row+1):
                if db.cell(row=prod, column=3).value != None and (
                    db.cell(row=prod, column=3).value not in competency[db.cell(row=16, column=user).value][db.cell(row=prod, column=2).value]):
                    competency[db.cell(row=16, column=user).value][db.cell(row=prod, column=2).value][db.cell(row=prod, column=3).value] = [
                                db.cell(row=prod, column=user).value, db.cell(row=prod, column=user+1).value
                            ]                        
    return competency[individual]

def filter_trainings(trainings, timezone='Asia/Kolkata'):
    """Filter trainings based on datetime.
    Args:
        trainings (object): List of trainings to filter.
    Returns:
        Dictionary of trainings filtered based on datetime.
    """
    _today = datetime.now(tzlocal.get_localzone())
    _timezone = pytz.timezone(timezone)
    returnable = {'completed': [], 'not_completed': []}
    for _training in trainings:
        _overall_time = _training.training_date + timedelta(hours=_training.training_duration)
        _overall_time = _overall_time.astimezone(_timezone)
        if _today >= _overall_time:
            returnable['completed'].append(_training)
        else:
            returnable['not_completed'].append(_training)
    return returnable

def get_last_modified_of_all_users(members):
    """Fetch last modified date of all users.

    Args:
        members (object): Members object from database.

    Returns:
        _last_modifies (dict) : Last modified date of all users.
    """
    _last_modifies = dict()
    for member in members:
        _name = member.first_name + ' ' + member.last_name
        _last_modifies[_name] = get_last_modified_time(_name)
    return _last_modifies

def get_last_modified_time(individual):
    """Fetch last updated time of user from excel
    
    Args:
        individual (str): Firstname + Lastname of user from database to fetch last modified date.

    Returns:
        date (str): Date time from the excel
    """
    database = openpyxl.load_workbook('Application/static/db.xlsx')
    db = database[WORKSHEET_NAME]

    for col in range(4, db.max_column+1):
        if db.cell(row=16, column=col).value == individual:
            _date = db.cell(row=16, column=(col+1)).value
            return str(_date)
    return ''

def update_modified_time(individual):
    """Update last updated time in IST in excel db
    
    Args:
        individual (str): Firstname + Lastname of user from database to fetch last modified date.
    """
    database = openpyxl.load_workbook('Application/static/db.xlsx')
    db = database[WORKSHEET_NAME]

    for col in range(4, db.max_column+1):
        if db.cell(row=16, column=col).value == individual:
            current_time = datetime.now(pytz.timezone('Asia/Kolkata')).strftime('%d %b %Y, %H:%M')
            db.cell(row=16, column=(col+1)).value = str(current_time)
    database.save('Application/static/db.xlsx')

def update_user_and_modules():
    """Fetch username, modules and products from database and update it in excel datbase (Location : 'static/db.xlsx')
    """
    database = openpyxl.load_workbook('Application/static/db.xlsx')
    db = database[WORKSHEET_NAME]

    row_start = 18
    column_start = 2 

    # Arranging modules and products in excel
    for mod in module.objects.all():
        for prod in product.objects.all():
            if prod.module == mod:
                db.cell(row=row_start, column=column_start).value = str(mod) # module name in 'B' column
                db.cell(row=row_start, column=column_start+1).value = str(prod) # product name in 'C' column
                row_start += 1
        db.cell(row=row_start, column=column_start+1).value = None
        row_start += 1

    # Adding users in excel
    user_row = 16
    user_col = 4

    for usr in User.objects.all():
        if str(usr) != 'admin': # neglecting admin user
            db.cell(row=user_row, column=user_col).value = str(usr.first_name) + ' ' + str(usr.last_name)
            db.cell(row=user_row+1, column=user_col).value = 'Current'
            db.cell(row=user_row+1, column=user_col+1).value = 'Target'
            user_col += 2

    database.save('Application/static/db.xlsx')

def delete_products(id):
    """Fetch product id  from database and update it in excel datbase after deleting respective product (Location : 'static/db.xlsx')
    Args:
    id (str): product_id
    """
    database = openpyxl.load_workbook('Application/static/db.xlsx')
    db = database[WORKSHEET_NAME]

    prod = product.objects.get(pk=id)
    for r in range(18, db.max_row + 1):
        if db.cell(row=r, column=3).value == str(prod):
            db.delete_rows(idx=r, amount=1)

    database.save('Application/static/db.xlsx')

def delete_modules(id):
    """Fetch module id  from database and update it in excel datbase after deleting respective module (Location : 'static/db.xlsx')
    Args:
    id (str): module_id
    """
    database = openpyxl.load_workbook('Application/static/db.xlsx')
    db = database[WORKSHEET_NAME]

    mod = module.objects.get(pk=id)
    for r in range(18, db.max_row + 1):
        if db.cell(row=r, column=2).value == str(mod):
            db.delete_rows(idx=r, amount=1)

    database.save('Application/static/db.xlsx')